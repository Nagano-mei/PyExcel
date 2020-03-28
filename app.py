import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
# print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
# cell = sheet.cell(row=1, column=1)

column = sheet["a"]
cells = sheet["a:c"]
# sheet["a1:c3"]
# sheet[1:3]
sheet.append([1004, 2, "$3"])
# sheet.insert_rows()
wb.save("transactions2.xlsx")

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column +1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet("Sheet1")